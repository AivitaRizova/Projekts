from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb_t = load_workbook('izejvielas_materiali_T.xlsx', data_only=True)
ws_t = wb_t.active

max_row = ws_t.max_row
# Cikls, kas aprēķina rindu summas
for row in range(9, 15): # Ņem vērtības no rindām 9.-15.
    row_sum = 0 #tiek piešķirta vērtība mainīgajam

    for col in range(3, 15): # Ņem vērtības no 3. līdz 14. kolonnai
        char = get_column_letter(col)
        cell_value = ws_t[char + str(row)].value # tiek iegūta šunas vērtība

        if cell_value is not None and isinstance(cell_value, (int, float)): # pārbauda vai ir vērtība un vai ir vesels skaitlis vai decimāl skaitlis(int, float)
            row_sum += cell_value
    ws_t['O' + str(row)] = row_sum # rezultāts tiek ievietos O kolonnā

for col in range(3, 15):
    col_sum = 0 # mainīgajam pieškir vērtību

    for row in range(9, 15):
        cell_value = ws_t[get_column_letter(col) + str(row)].value

        if cell_value is not None and isinstance(cell_value, (int, float)): #pārbauda vai ir vērtība un vai ir vesels skaitlis vai decimāl skaitlis(int, float)
            col_sum += cell_value
    ws_t[get_column_letter(col) + str(15)] = col_sum # saglabā summu 15.rindā tajā pašā kolonnā

col_sum_total = sum(ws_t[get_column_letter(col) + str(15)].value for col in range(3, 15)) # aprēķina summu no vairaku kolonnu 15.rindas šūnām, diapozonā no 3 līdz 14
ws_t['O15'] = round(col_sum_total, 2) # Summa tiek ievietota O kolonnas 15.rindas šūnā un rezultātu noapaļo līdz 2 cipariem aiz komata

wb_t.save('izejvielas_materiali_T.xlsx') # saglabā failā
# to pašu atārto ar otru failu
wb_k = load_workbook('izejvielas_materiali_K.xlsx', data_only=True)
ws_k = wb_k.active

max_row = ws_k.max_row
for row in range(9, 16): # rindas diapazona no 9 līdz 15
    row_sum = 0

    for col in range(3, 15):
        char = get_column_letter(col)
        cell_value = ws_k[char + str(row)].value

        if cell_value is not None and isinstance(cell_value, (int, float)):
            row_sum += cell_value
    ws_k['O' + str(row)] = row_sum

for col in range(3, 15):
    col_sum = 0

    for row in range(9, 16):
        cell_value = ws_k[get_column_letter(col) + str(row)].value

        if cell_value is not None and isinstance(cell_value, (int, float)):
            col_sum += cell_value
    ws_k[get_column_letter(col) + str(16)] = col_sum # tiek saglabāts rezultāts 15 rindā

col_sum_total = sum(ws_k[get_column_letter(col) + str(16)].value for col in range(3, 15))
ws_k['O16'] = round(col_sum_total, 2)

wb_k.save('izejvielas_materiali_K.xlsx')

value_t = ws_t['O15'].value # iegūst vērtību no O15 un saglabā mainīgajā
value_k = ws_k['O16'].value # iegūst vērtību no O16 un saglabā mainīgajā

if value_t > value_k: # salīdzina vērtības un izdrukā atbilstošo paziņojumu par lielākajiem ienākumiem.
    print("Lielākie ienākumi ir: izejvielas_materiali_T")
elif value_k > value_t:
     print("Lielākie ienākumi ir: izejvielas_materiali_K")
else:
    print("Abos ir vienādi ienākumi.")

wb_t.close()
wb_k.close()